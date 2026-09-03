"""@brief Viewer for tabular data: CSV/TSV and Excel .xlsx files.

@details CSV files are parsed with the csv module (delimiter sniffing);
XLSX files are read sheet-by-sheet with openpyxl. Rows are shown in a
QTableWidget with generous caps to keep the UI responsive; truncation is
announced with a banner. Everything is read in-process.
"""

import csv

from PyQt6.QtWidgets import (
    QComboBox,
    QHeaderView,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from PyQt6.QtGui import QFont

from universal_viewer.filetypes import get_extension
from universal_viewer.utils import read_text_preview
from universal_viewer.viewers.base import BaseViewer

#: @brief Maximum number of rows loaded into the table widget.
_MAX_ROWS = 5000

#: @brief Maximum number of columns loaded into the table widget.
_MAX_COLS = 256

#: @brief Bytes sampled from the file start for delimiter sniffing.
_SNIFF_SAMPLE = 64 * 1024

#: @brief Extensions handled by the table viewer.
_TABLE_EXTENSIONS = frozenset({".csv", ".tsv", ".xlsx", ".xlsm"})


class TableViewer(BaseViewer):
    """@brief Spreadsheet-style viewer for CSV and XLSX files."""

    SUPPORTED_EXTENSIONS = _TABLE_EXTENSIONS

    def __init__(self, parent: QWidget | None = None) -> None:
        """@brief Construct the widget with its sheet selector and grid.

        @param parent: Optional parent widget.
        """
        super().__init__(parent)
        self._path = ""
        self._headers: list[str] = []
        self._rows: list[list[str]] = []
        self._truncated = False

        self._sheet_combo = QComboBox(self)
        self._sheet_combo.setVisible(False)
        self._sheet_combo.currentTextChanged.connect(self._on_sheet_changed)

        self._banner = QLabel(self)
        self._banner.setStyleSheet("background: #fff3cd; color: #664d03; padding: 4px;")
        self._banner.hide()

        self._table = QTableWidget(self)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        header = self._table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._table.verticalHeader().setVisible(True)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addWidget(self._sheet_combo)
        layout.addWidget(self._banner)
        layout.addWidget(self._table, stretch=1)

    def load(self, path: str) -> None:
        """@brief Load a CSV or XLSX file into the grid.

        @param path: Absolute path of the table file to display.
        """
        self._path = path
        extension = get_extension(path)
        if extension in (".xlsx", ".xlsm"):
            self._load_xlsx()
        else:
            self._load_csv()

    def _load_csv(self) -> None:
        """@brief Parse the current CSV/TSV file and fill the grid.

        @details The dialect is sniffed from the file head; encoding is
        detected automatically.
        """
        try:
            text, _encoding, _truncated = read_text_preview(self._path)
        except OSError as error:
            self._show_error(f"Could not read file: {error}")
            return
        lines = text.splitlines()
        sample = "\n".join(lines[:200])
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        try:
            rows = list(csv.reader(lines, dialect))
        except csv.Error as error:
            self._show_error(f"CSV parsing failed: {error}")
            return
        self._headers = [f"Col {index + 1}" for index in range(len(rows[0]) if rows else 0)]
        self._truncated = len(rows) > _MAX_ROWS
        self._rows = rows
        self._sheet_combo.setVisible(False)
        self._fill_grid()

    def _load_xlsx(self) -> None:
        """@brief Read the current XLSX workbook and populate the sheet list.

        @details The first sheet is shown initially; switching sheets rebuilds
        the grid from the cached workbook handle.
        """
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        except ImportError:
            self._show_error("openpyxl package is not installed.")
            return
        except Exception as error:
            self._show_error(f"Could not open workbook: {error}")
            return
        sheet_names = list(workbook.sheetnames)
        if not sheet_names:
            workbook.close()
            self._show_error("Workbook has no sheets.")
            return
        sheet = workbook[sheet_names[0]]
        rows: list[list[str]] = []
        truncated = False
        try:
            for row in sheet.iter_rows(values_only=True):
                rows.append(["" if cell is None else str(cell) for cell in row])
                if len(rows) >= _MAX_ROWS:
                    truncated = True
                    break
        finally:
            workbook.close()
        width = max((len(row) for row in rows), default=0)
        self._headers = [f"Col {index + 1}" for index in range(width)]
        self._rows = rows
        self._truncated = truncated
        self._sheet_combo.blockSignals(True)
        self._sheet_combo.clear()
        self._sheet_combo.addItems(sheet_names)
        self._sheet_combo.setCurrentIndex(0)
        self._sheet_combo.blockSignals(False)
        self._sheet_combo.setVisible(len(sheet_names) > 1)
        self._fill_grid()

    def _fill_grid(self) -> None:
        """@brief Populate the QTableWidget from the cached rows.

        @details Caps the number of rows/columns and shows a banner when
        truncated; the first parsed row is rendered bold.
        """
        rows = self._rows[:_MAX_ROWS]
        width = min(max((len(row) for row in rows), default=0), _MAX_COLS)
        self._table.setRowCount(len(rows))
        self._table.setColumnCount(width)
        self._table.setHorizontalHeaderLabels(self._headers[:width])
        bold = QFont()
        bold.setBold(True)
        for row_index, row in enumerate(rows):
            for col_index in range(width):
                value = row[col_index] if col_index < len(row) else ""
                item = QTableWidgetItem(value)
                if row_index == 0:
                    item.setFont(bold)
                self._table.setItem(row_index, col_index, item)
        self._table.resizeColumnsToContents()
        for column in range(width):
            if self._table.columnWidth(column) > 300:
                self._table.setColumnWidth(column, 300)
        if self._truncated:
            self._banner.setText(f"Showing the first {_MAX_ROWS} rows.")
            self._banner.show()
        elif width == _MAX_COLS:
            self._banner.setText(f"Showing the first {_MAX_COLS} columns.")
            self._banner.show()
        else:
            self._banner.hide()

    def _on_sheet_changed(self, sheet_name: str) -> None:
        """@brief Slot: rebuild the grid when another sheet is selected.

        @details On failure the error is reported through the status message
        instead of being swallowed, and the previously shown data is kept.

        @param sheet_name: Name of the newly selected worksheet.
        """
        if not sheet_name or not self._path.lower().endswith((".xlsx", ".xlsm")):
            return
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]
            rows: list[list[str]] = []
            truncated = False
            try:
                for row in sheet.iter_rows(values_only=True):
                    rows.append(["" if cell is None else str(cell) for cell in row])
                    if len(rows) >= _MAX_ROWS:
                        truncated = True
                        break
            finally:
                workbook.close()
            width = max((len(row) for row in rows), default=0)
            self._headers = [f"Col {index + 1}" for index in range(width)]
            self._rows = rows
            self._truncated = truncated
            self._fill_grid()
        except Exception as error:
            self.status_message.emit(f"Could not switch sheet: {error}")

    def _show_error(self, message: str) -> None:
        """@brief Show an error banner instead of table data.

        @param message: Human-readable problem description.
        """
        self._table.setRowCount(0)
        self._table.setColumnCount(1)
        self._banner.setText(message)
        self._banner.show()
        self.status_message.emit(message)

    def cleanup(self) -> None:
        """@brief Drop the cached rows."""
        self._rows = []
        self._truncated = False
        self._table.setRowCount(0)
